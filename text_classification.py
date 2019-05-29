import openpyxl

filename = '贪污受贿罪_一审_清洗后2.xlsx'

file = openpyxl.load_workbook(filename)

ws = file.worksheets[0]
print("open load_file")
file_write = openpyxl.Workbook()
write = file_write.create_sheet(index=0)  # 创建新表
file_write_2 = openpyxl.Workbook()
write2 = file_write_2.create_sheet(index=0)
print("open write_file")
row = ws.max_row
column = ws.max_column
print(row, column)
o = 1
p = 1
for i in range(1, column + 1):
    cell_value = ws.cell(row=o, column=i).value
    write.cell(row=o, column=i).value = cell_value
    write2.cell(row=p, column=i).value = cell_value
o += 1
p += 1
for i in range(2, row + 1):
    print(i)
    value = ws.cell(row=i, column=5).value
    print(value)
    judge = False
    if value == "贪污罪":
        judge = True
    print(judge)
    if judge:
        for j in range(1, column + 1):
            cell_value = ws.cell(row=i, column=j).value
            write.cell(row=o, column=j).value = cell_value
        o += 1
    if not judge:
        for j in range(1, column + 1):
            cell_value = ws.cell(row=i, column=j).value
            write2.cell(row=p, column=j).value = cell_value
        p += 1

file_write.save(filename="纯贪污罪_一审_清洗.xlsx")
file_write_2.save(filename='非纯贪污罪_一审_清洗.xlsx')
