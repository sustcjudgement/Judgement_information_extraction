import openpyxl
import re
import matplotlib.pyplot as plt
import echarts_china_provinces_pypkg
from pyecharts import options as opts
from pyecharts.charts import Geo,Radar,Bar
from pyecharts.globals import ChartType, SymbolType

plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号


# 输入聚类结果文件 ，所分的类别数目
# 输出 每个类别 的案件编号 dict（list）
def take_part(filename, part):
    part_dict = dict()
    for i in range(part):
        part_dict[i] = list()
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    row = ws.max_row
    for i in range(1, row):
        par = int(ws.cell(row=i, column=4).value)  # 所分类别在第四列
        point = int(ws.cell(row=i, column=1).value)  # 所属案件编号在第一列
        part_dict[par].append(point)
    return part_dict


# 输入 1.法条抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个法条在该类下的频率
def read_one_part_law(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    law_list = ['382t', '383t', '385t', '386t', '387t', '388t', '394t', ]
    law_dict = dict()
    for i in law_list:
        law_dict[i] = 0
    for i in part_dict[part]:
        law = str(ws.cell(row=i - 1, column=2).value).replace(';', '').split(' ')
        for t in law_list:
            if t in law:
                law_dict[t] += 1
    return law_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个地区在该类下的频率
def read_one_part_local(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    local_dict = dict()
    for i in part_dict[part]:
        local = str(ws.cell(row=i, column=1).value)
        if local in local_dict:
            local_dict[local] += 1
        else:
            local_dict[local] = 1
    return local_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个职位在该类下的频率
def read_one_part_work(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    work_dict = dict()
    for i in part_dict[part]:
        work = str(ws.cell(row=i, column=2).value)
        if work != 'None':
            work = work[-2:]
            if work in work_dict:
                work_dict[work] += 1
            else:
                work_dict[work] = 1

    return work_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个年龄段在该类下的频率
def read_one_part_old(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    old_list = [20, 30, 40, 50, 60]
    old_dict = dict()
    for i in old_list:
        old_dict[i] = 0
    for i in part_dict[part]:
        old = int(ws.cell(row=i, column=4).value)
        for t in old_list:
            if old < t + 10 and old >= t:
                old_dict[t] += 1
    return old_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个作案时长段在该类下的频率
def read_one_part_time(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    time_list = [1, 6, 11, 16, 21, 26, 31, 36]
    time_dict = dict()
    for i in time_list:
        time_dict[i] = 0
    for i in part_dict[part]:
        time_part = str(ws.cell(row=i, column=5).value).split('-')
        time = int(time_part[0]) - int(time_part[1])
        for t in time_list:
            if time < t + 5 and time >= t:
                time_dict[t] += 1
    return time_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个作案时长段在该类下的频率
def read_one_part_money(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    money_list = ['10000-30000', '30000-200000', '200000-3000000', '3000000-99999999999']
    money_dict = dict()
    for i in money_list:
        money_dict[i] = 0
    for i in part_dict[part]:
        money_part = str(ws.cell(row=i, column=6).value)
        money = 0
        try:
            if money_part[-2:] == '万元':
                money = int(money_part.split('.')[0]) * 10000
            else:
                money = int(money_part.split('.')[0])
            for t in money_list:
                money_arr = t.split('-')
                money_low = int(money_arr[0])
                money_high = int(money_arr[1])
                if money < money_high and money >= money_low:
                    money_dict[t] += 1
        except:
            pass
    return money_dict
# 输入 1.学历抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每种学历在该类下的频率
def read_one_part_study(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    study_list = ['硕士|研究生', '初中', '高中', '大学|本科', '大专|专科', '中专', '无','小学' ]
    low_dict = dict()
    for i in study_list:
        low_dict[i] = 0
    for i in part_dict[part]:
        low = str(ws.cell(row=i - 1, column=3).value)
        for t in study_list:
            if low in t:
                low_dict[t] += 1
    return low_dict
# 输入 1.刑期抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个刑期长段在该类下的频率
def read_one_part_push_data(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    time_list = [0, 10, 20, 30, 40, 50, 60, 70,80,90,100,110,120,130,140,150,160,170,180]
    time_dict = dict()
    time_dict['无期'] = 0
    for i in time_list:
        time_dict[i] = 0
    for i in part_dict[part]:
        time = str(ws.cell(row=i, column=8).value)
        if time != 'None':
            time = int(time)
            for t in time_list:
                if time < t + 10 and time >= t:
                    time_dict[t] += 1
        else:
            time_dict['无期']  += 1
    return time_dict

# 输入list 每个分类的dict存储数据
# 输出图片
def draw_bar(list_dic, word,int_flag = False) -> Bar:
    point = 0
    ret_list = []
    # print(list_dic)
    for i in list_dic:
        x = list()
        y = list()
        for t in i:
            y.append(i[t])
            t = t.split("|")[0]
            x.append(t)
        # plt.bar(range(len(y)), y, color='rgb', tick_label=x)
        save_place = './result/' + word + str(point) + '.jpg'
        point += 1
        plt.savefig(save_place)
        plt.show()
        # ============test=============
        # for i in range(len(x)):
        #     if x[i] is "广西壮族自治":
        #         x[i] = "广西"
        # ===============graph================
        # c = (
        #     Geo()
        #         .add_schema(maptype="china")
        #         .add("geo", [list(z) for z in zip(x, y)])
        #         .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
        #         .set_global_opts(
        #         visualmap_opts=opts.VisualMapOpts(is_piecewise=True),
        #         title_opts=opts.TitleOpts(title="Geo-基本示例"),
        #     )
        # )
        c = (
            Bar()
                .add_xaxis(x)
                .add_yaxis("被告人职位",y)
                .set_global_opts(title_opts=opts.TitleOpts(title="Cluster"+str(point)),
                                 yaxis_opts=opts.AxisOpts(name="个数"),
                                 )
        )
        ret_list.append(c)
        # c.render()
        print("oh yes")
    return ret_list

def main(filename_1,filename_2,part,name):
    part_dict = take_part(filename_2, part)
    law_list = list()  # 法条
    local_list = list()  # 案发地
    work_list = list()  # 职位
    old_list = list()  # 年龄
    time_list = list()  # 涉案时长
    money_list = list()  # 涉案金额
    push_list = list()   #刑期
    study_list = list()  #学历
    # add 刑期 学历 画图
    #==================================================================
    # for x in range(part):
    #     push_dict = read_one_part_push_data(filename_3, part_dict, x)
    #     push_list.append(push_dict)
    # draw_bar(push_list, 'punish')
    # ret = draw_bar(push_list, 'punish' + name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="punish"+str(j)+".html")
    #     j -= 1
    #     print(j)
    # ==================================================
    # for x in range(part):
    #     study_dict = read_one_part_study(filename_3, part_dict, x)
    #     study_list.append(study_dict)
    # # draw_bar(study_list, 'study')
    # ret = draw_bar(study_list, 'study' + name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="study"+str(j)+".html")
    #     j -= 1
    #     print(j)
    # ===========================================
    # for x in range(part):
    #     law_dict = read_one_part_law(filename_2, part_dict, x)
    #     law_list.append(law_dict)
    # # draw_bar(law_list, 'law'+name)
    # ret = draw_bar(law_list, 'law' + name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="law"+str(j)+".html")
    #     j -= 1
    #     print(j)
    # ========================================================================================================================
    # for x in range(part):
    #     local_dict = read_one_part_local(filename_1, part_dict, x)
    #     local_dict_sim = dict()
    #     for i in local_dict:
    #         if local_dict[i] > 50:
    #             local_dict_sim[i] = local_dict[i]
    #     local_list.append(local_dict_sim)
    # ret = draw_bar(local_list, 'local'+name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="local"+str(j)+".html")
    #     j-=1
    # ========================================================
    for x in range(part):
        work_dict = read_one_part_work(filename_1, part_dict, x)
        work_dict_sim = dict()
        for i in work_dict:
            if work_dict[i] > 5:
                work_dict_sim[i] = work_dict[i]
        work_list.append(work_dict_sim)
    ret = draw_bar(work_list, 'job' + name)
    j = 3
    for i in range(len(ret)):
        ret.pop().render(path="job"+str(j)+".html")
        j-=1
    # # ========================================================================================================================
    # for x in range(part):
    #     old_dict = read_one_part_old(filename_1, part_dict, x)
    #     old_list.append(old_dict)
    # # draw_bar(old_list, 'old'+name,int_flag=True)
    # ret = draw_bar(old_list, 'old' + name)
    # print(len(ret))
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="old"+str(j)+".html")
    #     j-=1
    # # ========================================================================================================================
    # for x in range(part):
    #     time_dict = read_one_part_time(filename_1, part_dict, x)
    #     time_list.append(time_dict)
    # draw_bar(time_list, 'time'+name,int_flag=True)
    # ret = draw_bar(time_list, 'time' + name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="time"+str(j)+".html")
    #     j -= 1
    #     print(j)
    # # ========================================================================================================================
    # for x in range(part):
    #     money_dict = read_one_part_money(filename_1, part_dict, x)
    #     money_list.append(money_dict)
    # # draw_bar(money_list, 'money'+name)
    # ret = draw_bar(money_list, 'money' + name)
    # j = 3
    # for i in range(len(ret)):
    #     ret.pop().render(path="money"+str(j)+".html")
    #     j -= 1
    #     print(j)

if __name__ == '__main__':
    filename_1 = r'E:\data_analysis\Judgement_information_extraction\database_xml\基本信息提取1.xlsx'  # 基本信息提取
    filename_2 = r'E:\data_analysis\Judgement_information_extraction\database_xml\贪污受贿罪_一审_提取_birch.xlsx'  # 聚类结果
    filename_3 = r'E:\data_analysis\Judgement_information_extraction\database_xml\抽取.xlsx'  # 聚类结果
    part = 4
    name = "_birch_"
    main(filename_1 , filename_2,part,name)
