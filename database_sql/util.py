
from openpyxl import load_workbook
from numpy import *
from database.connect_database import *

def convert():
    pe = load_workbook(r"C:\Users\Vold\Desktop\base.xlsx")  # 默认可读写，若有需要可以指定write_only和read_only为True
    sheet = pe.worksheets[1]
    print("load database is correct")
    length = len(sheet["B"])
    for i in range(2, length + 2):
        print(str(sheet.cell(i, 1).value))
        pb = AgeBean(
            # judgement_id=i,
            #                area=str(sheet.cell(i, 1).value),
            #                job=str(sheet.cell(i, 2).value),
                           age=str(sheet.cell(i, 4).value),
            #                time=str(sheet.cell(i, 5).value),
            #                money=str(sheet.cell(i, 6).value)
                           )
        insert_DB(pb)

if __name__ == '__main__':

    convert()