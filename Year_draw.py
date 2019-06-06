import openpyxl
import re
import matplotlib.pyplot as plt
import echarts_china_provinces_pypkg
from pyecharts import options as opts
from pyecharts.charts import Geo,Radar,Bar
from pyecharts.globals import ChartType, SymbolType

plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号


year = []
year = list(year)
# 输入聚类结果文件 ，所分的类别数目
# 输出 每个类别 的案件编号 dict（list）
def take_part(filename, part):
    part_dict = dict()
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    row = ws.max_row
    year = set()
    for i in range(2, row+1):
        if ws.cell(row=i, column=11).value is not None:
            par = int(ws.cell(row=i, column=11).value)  # 年份在第四列
            point = i  # 所属案件编号在第一列
            if part_dict.get(par,"n") is "n":
                part_dict[par] = list()
            part_dict[par].append(point)
            # print(i,"????",par)
            year.add(par)
    return part_dict,year

def read_one_part_old(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    old_list = [20, 30, 40, 50, 60]
    old_dict = dict()
    for i in old_list:
        old_dict[i] = 0
    for i in part_dict[part]:
        old = int(ws.cell(row=i, column=4).value)
        for t in old_list:
            if old < t + 10 and old >= t:
                old_dict[t] += 1
    return old_dict

# 输入 1.法条抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个法条在该类下的频率
def read_one_part_law(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    law_list = ["382t", '383t', '385t', '386t', '387t', '388t', '394t', ]
    law_dict = dict()
    for i in law_list:
        law_dict[i] = 0
    for i in part_dict[part]:
        law = str((ws.cell(row=i - 1, column=9).value).replace(';', '').split(' '))
        # print(law)
        for t in law_list:
            if t in law:
                law_dict[t] += 1
    print(law_dict)
    return law_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个地区在该类下的频率
def read_one_part_local(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    local_dict = dict()
    for i in part_dict[part]:
        local = str(ws.cell(row=i, column=2).value)
        if local in local_dict:
            local_dict[local] += 1
        else:
            local_dict[local] = 1
    return local_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个职位在该类下的频率
def read_one_part_work(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    work_dict = dict()
    if part is 2000:
        print("???????")
    for i in part_dict[part]:
        work = str(ws.cell(row=i, column=4).value)
        if work != 'None':
            if work in work_dict:
                work_dict[work] += 1
            else:
                work_dict[work] = 1

    return work_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个作案时长段在该类下的频率
def read_one_part_time(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    time_list = [1, 6, 11, 16, 21, 26, 31, 36]
    time_dict = dict()
    for i in time_list:
        time_dict[i] = 0
    for i in part_dict[part]:
        time_part = str(ws.cell(row=i, column=5).value).split('-')
        time = int(time_part[0]) - int(time_part[1])
        for t in time_list:
            if time < t + 5 and time >= t:
                time_dict[t] += 1
    return time_dict


# 输入 1.信息抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个作案时长段在该类下的频率
def read_one_part_money(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    money_list = ['10000-30000', '30000-200000', '200000-3000000', '3000000-99999999999']
    money_dict = dict()
    for i in money_list:
        money_dict[i] = 0
    for i in part_dict[part]:
        money_part = str(ws.cell(row=i, column=6).value)
        money = 0
        try:
            if money_part[-2:] == '万元':
                money = int(money_part.split('.')[0]) * 10000
            else:
                money = int(money_part.split('.')[0])
            for t in money_list:
                money_arr = t.split('-')
                money_low = int(money_arr[0])
                money_high = int(money_arr[1])
                if money < money_high and money >= money_low:
                    money_dict[t] += 1
        except:
            pass
    return money_dict
# 输入 1.学历抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每种学历在该类下的频率
def read_one_part_study(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    study_list = ['硕士|研究生', '初中', '高中', '大学|本科', '大专|专科', '中专', '无','小学' ]
    low_dict = dict()
    for i in study_list:
        low_dict[i] = 0
    for i in part_dict[part]:
        low = str(ws.cell(row=i - 1, column=3).value)
        for t in study_list:
            if low in t:
                low_dict[t] += 1
    return low_dict
# 输入 1.刑期抽取结果 2.每个类别 的案件编号 dict（list） 3.统计的那一类分类结果
# 输出 每个刑期长段在该类下的频率
def read_one_part_push_data(filename, part_dict, part):
    file = openpyxl.load_workbook(filename)
    ws = file.worksheets[0]
    time_list = [0, 10, 20, 30, 40, 50, 60, 70,80,90,100,110,120,130,140,150,160,170,180]
    time_dict = dict()
    time_dict['无期'] = 0
    for i in time_list:
        time_dict[i] = 0
    for i in part_dict[part]:
        time = str(ws.cell(row=i, column=8).value)
        if time != 'None':
            time = int(time)
            for t in time_list:
                if time < t + 10 and time >= t:
                    time_dict[t] += 1
        else:
            time_dict['无期']  += 1
    return time_dict

# 输入list 每个分类的dict存储数据
# 输出图片
def draw_bar(list_dic, word,year) -> Geo:
    point = 0
    ret_list = []
    print(list_dic)
    for i in list_dic:
        x = list()
        y = list()
        for t in i:
            # print("yes")
            y.append(i[t])
            # t = t.split("|")[0]
            # print(t)
            # if t == "广西壮族自治区":
            #     # print("hsdlkfhjwiefjwpoafp")
            #     t = "广西"
            # elif t == "新疆维吾尔自治区":
            #     t = "新疆"
            # elif t == "内蒙古自治区":
            #     t = "内蒙古"
            # elif t == "安徽省":
            #     t = "安徽"
            # elif t == "宁夏回族自治区":
            #     t = "宁夏"
            # else:
            #     t = t[:-1]
            x.append(t)
        # plt.bar(range(len(y)), y, color='rgb', tick_label=x)
        # save_place = './result/' + word + str(point) + '.jpg'
        point += 1
        # plt.savefig(save_place)
        # plt.show()
        # print(list(zip(x,y)))
        # print(x)
        # print(y)

        if len(x) is 0 or len(y) is 0:
            print(point-1)
            continue
        m = zip(x,y)
        year = list(year)
        # ===============graph================
        # c = (
        #     Geo()
        #         .add_schema(maptype="china")
        #         .add("geo", [list(z) for z in m])
        #         .set_series_opts(label_opts=opts.LabelOpts(is_show=False))
        #         .set_global_opts(
        #         visualmap_opts=opts.VisualMapOpts(is_piecewise=True),
        #         title_opts=opts.TitleOpts(title=year[point-1]),
        #     )
        # )

        c = (
            Bar()
                .add_xaxis(x)
                .add_yaxis("涉案金额",y)
                .set_global_opts(title_opts=opts.TitleOpts(title=year[point-1]),
                                 yaxis_opts=opts.AxisOpts(name="个数"),
                                 # xaxis_opts=opts.AxisOpts(name="t=条")
                                 )
        )
        ret_list.append(c)
        print("oh yes")
    return ret_list


def main(filename_1,filename_2,part,name):
    part_dict , year = take_part(filename_2, part)

    law_list = list()  # 法条
    local_list = list()  # 案发地
    work_list = list()  # 职位
    old_list = list()  # 年龄
    time_list = list()  # 涉案时长
    money_list = list()  # 涉案金额
    push_list = list()   #刑期
    study_list = list()  #学历
    # add 刑期 学历 画图
    #==================================================================
    # for x in year:
    #     push_dict = read_one_part_push_data(filename_3, part_dict, x)
    #     push_list.append(push_dict)
    # # draw_bar(push_list, 'punish')
    # year = list(year)
    # ret = draw_bar(push_list, 'punish' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     p = str(year[j])
    #     ret.pop().render(path="punish"+p+".html")
    #     j-=1
    # # ==================================================
    # for x in year:
    #     study_dict = read_one_part_study(filename_3, part_dict, x)
    #     study_list.append(study_dict)
    # # draw_bar(study_list, 'study')
    # ret = draw_bar(study_list, 'study' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="study"+p+".html")
    #     j-=1
    # # ===========================================
    # for x in year:
    #     law_dict = read_one_part_law(filename_2, part_dict, x)
    #     law_list.append(law_dict)
    # # # draw_bar(law_list, 'law'+name)
    # ret = draw_bar(law_list, 'law' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="law"+p+".html")
    #     j-=1
    # # ========================================================================================================================
    # for x in year:
    #     local_dict = read_one_part_local(filename_1, part_dict, x)
    #     local_dict_sim = dict()
    #     for i in local_dict:
    #         if local_dict[i] > 8:
    #             local_dict_sim[i] = local_dict[i]
    #     local_list.append(local_dict_sim)
    # ret = draw_bar(local_list, 'local' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="local"+p+".html")
    #     j-=1
    # # ========================================================================================================================
    # for x in year:
    #     work_dict = read_one_part_work(filename_1, part_dict, x)
    #     work_dict_sim = dict()
    #     for i in work_dict:
    #         if work_dict[i] > 5:
    #             work_dict_sim[i] = work_dict[i]
    #     work_list.append(work_dict_sim)
    # ret = draw_bar(work_list, 'job' + name)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="job"+p+".html")
    #     j-=1
    # # ========================================================
    #
    # for x in year:
    #     work_dict = read_one_part_work(filename_1, part_dict, x)
    #     work_dict_sim = dict()
    #     for i in work_dict:
    #         work_dict_sim[i] = work_dict[i]
    #     work_list.append(work_dict_sim)
    #     year.add(x)
    # year = list(year)
    # ret = draw_bar(work_list, 'job' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     p = str(year[j])
    #     ret.pop().render(path="job"+p+".html")
    #     j-=1
    # # ========================================================================================================================
    # for x in year:
    #     time_dict = read_one_part_time(filename_1, part_dict, x)
    #     time_list.append(time_dict)
    # # draw_bar(time_list, 'time'+name,int_flag=True)
    # ret = draw_bar(time_list, 'time' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="time"+p+".html")
    #     j-=1
    # # ========================================================================================================================
    # for x in year:
    #     money_dict = read_one_part_money(filename_1, part_dict, x)
    #     money_list.append(money_dict)
    # # # draw_bar(money_list, 'money'+name)
    # ret = draw_bar(money_list, 'money' + name,year)
    # j = 13
    # for i in range(len(ret)):
    #     year = list(year)
    #     p = str(year[j])
    #     ret.pop().render(path="money"+p+".html")
    #     j-=1
    # =================================================================================
    for x in range(part):
        old_dict = read_one_part_old(filename_1, part_dict, x)
        old_list.append(old_dict)
    # draw_bar(old_list, 'old'+name,int_flag=True)
    ret = draw_bar(old_list, 'old' + name,year)
    j = 13
    for i in range(len(ret)):
        year = list(year)
        p = str(year[j])
        ret.pop().render(path="old"+p+".html")
        j-=1
if __name__ == '__main__':
    filename_1 = r'E:\data_analysis\Judgement_information_extraction\database_xml\抽取2.xlsx'  # 基本信息提取
    filename_2 = r'E:\data_analysis\Judgement_information_extraction\database_xml\抽取2.xlsx'  # 聚类结果
    filename_3 = r'E:\data_analysis\Judgement_information_extraction\database_xml\抽取2.xlsx'  # 聚类结果
    part = 4
    name = "_birch_"
    main(filename_1 , filename_2,part,name)
    print("over")