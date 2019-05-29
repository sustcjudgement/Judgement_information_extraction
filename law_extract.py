from openpyxl import load_workbook as lw
import re
import pandas as pd
import openpyxl
import numpy as np

end = ['判决', '如下', '判处']  # 读取的截止词
law_num = 29
LAW = [['382t'], ['383t'], ['385t'], ['386t'], ['387t'], ['388t'], ['394t'], ['382t', '1k'], ['382t', '2k'],
       ['382t', '3k'],
       ['383t', '1k'], ['383t', '2k'], ['383t', '3k'], ['383t', '4k'], ['383t', '1x'], ['383t', '2x'], ['383t', '3x'],
       ['383t', '1k', '1x'], ['383t', '1k', '2x'], ['383t', '1k', '3x'], ['385t', '1k'], ['385t', '2k'], ['386t', '1k'],
       ['387t', '1k'], ['387t', '2k'],
       ['388t', '1k'], ['388t', '2k'], ['388t', '3k'], ['394t', '1k']]

law = {'382t': [['382t'], ['382t', '1k'], ['382t', '2k'], ['382t', '3k']],
       '383t': [['383t'], ['383t', '1k'], ['383t', '2k'], ['383t', '3k'], ['383t', '4k'], ['383t', '1x'],
                ['383t', '2x'], ['383t', '3x'], ['383t', '1k', '1x'], ['383t', '1k', '2x'], ['383t', '1k', '3x']],
       '385t': [['385t'], ['385t', '1k'], ['385t', '2k']],
       '386t': [['386t'], ['386t', '1k']],
       '387t': [['387t'], ['387t', '1k'], ['387t', '2k']],
       '388t': [['388t'], ['388t', '1k'], ['388t', '2k'], ['388t', '3k']],
       '394t': [['394t'], ['394t', '1k']]
       }  # 匹配的法律条款


# 以下四个函数用于从庭审结果文本中提取出现的法条并将其格式化后写入数据库中

# FUNCTION:catch(judgment_text)
# 一、作用：捕捉结果关键字
# 二、函数输入及其默认值：
# 1.judgement_text:一个审判结果文本
# 三、函数输出及其默认值：
# 1. fa_list
def catch(judgment_text):
    # 法条集合--[（）,(法律，具体信息)，，，，]
    law_list = []
    i = 0
    while i < len(judgment_text):
        # 书名号开始记录
        if judgment_text[i] == '《':
            # 下一个字符
            j = i + 1
            # 这一个法条的法律
            fa = []
            # 直到书名号结束
            while judgment_text[j] != '》':
                fa.append(judgment_text[j])
                j += 1
            # 获得该法条法律
            fa = ''.join(fa)
            # 获得法条具体信息
            more_infor = []
            j += 1
            # 继续标志
            # flag = 0
            while judgment_text[j] != '《' and ''.join(judgment_text[j:j + 2]) not in end:
                # 解决多个法律的问题
                # if judgment_text[j] == '《':
                #     flag = 1
                #     break
                # else:
                more_infor.append(judgment_text[j])
                j += 1
            more_infor = ''.join(more_infor)
            # 将（法律，具体信息）二元组加入law_list
            law_list.append((fa, more_infor))
            # 继续读取
            # if flag == 0:
            #     i = j + 1
            # else:
            #     i = j
            if judgment_text[j] == '《':
                i = j
            if ''.join(judgment_text[j:j + 2]) in end:
                return law_list
        # 到“判决如下”截至
        elif ''.join(judgment_text[i:i + 2]) in end:
            break
        else:
            i += 1
    return law_list


# FUNCTION:divide(fa_list)
# 一、作用：格式化表述法条
# 二、函数输入及其默认值：
# 1.fa_list:一个审判结果文本:[('中华人民共和国刑法', '第三百八十二条第二款、第三百八十三条第一款第（三）项、第六十七条第一款、第七十二条第一款、第七十三条第二款、第三款的规定')]
# 三、函数输出及其默认值：
# 1.all_infor :[('中华人民共和国刑法', [['382t', '2k'], ['383t', '1k', '3x'], ['67t', '1k'], ['72t', '1k'], ['73t', '2k'], ['73t', '3k']])]
def divide(fa_list):
    # 获得格式化后的法条
    all_infor = []
    for i in fa_list:
        infor = []
        fa = i[0]
        if (re.search("刑法", fa) == None and re.search("刑事诉讼", fa) == None) or re.search("关于", fa) != None:
            continue
        more_infor = i[1]
        more_infor = more_infor.replace('，', '、')
        more_infor = more_infor.replace(',', '、')
        more_infor = more_infor.replace('；', '、')
        more_infor = more_infor.replace('和', '、')
        more_infor = more_infor.replace('以及', '、')
        status = more_infor.split('、第')
        # j是每个小法条
        for j in status:
            number_list = []
            # 维护一个头head标记
            head = 0
            for num in range(0, len(j)):
                if j[num] == '条':
                    this_num = []
                    for c in range(0, len(j[head:num])):
                        if str(word_to_number(j[head:num][c])) != '':
                            if word_to_number(j[head:num][c]) != 0:
                                this_num.append(str(word_to_number(j[head:num][c])))
                            else:
                                if c == 0:
                                    this_num.append(str(1))
                                elif str(word_to_number(j[head:num][c - 1])) == '':
                                    this_num.append(str(1))

                    this_num = ''.join(this_num) + 't'
                    number_list.append(this_num)
                    head = num
                elif j[num] == '款':
                    this_num = []
                    for c in range(0, len(j[head:num])):
                        if str(word_to_number(j[head:num][c])) != '':
                            if word_to_number(j[head:num][c]) != 0:
                                this_num.append(str(word_to_number(j[head:num][c])))
                            else:
                                if c == 0:
                                    this_num.append(str(1))
                                elif str(word_to_number(j[head:num][c - 1])) == '':
                                    this_num.append(str(1))
                    this_num = ''.join(this_num) + 'k'
                    number_list.append(this_num)
                    head = num
                elif j[num] == '项':
                    this_num = []
                    for c in range(0, len(j[head:num])):
                        if str(word_to_number(j[head:num][c])) != '':
                            if word_to_number(j[head:num][c]) != 0:
                                this_num.append(str(word_to_number(j[head:num][c])))
                            else:
                                if c == 0:
                                    this_num.append(str(1))
                                elif str(word_to_number(j[head:num][c - 1])) == '':
                                    this_num.append(str(1))
                    this_num = ''.join(this_num) + 'x'
                    number_list.append(this_num)
                    head = num
            infor.append(number_list)
        for m in range(0, len(infor)):
            n = infor[m]
            if n == []:
                pass
            elif 't' not in n[0]:
                if len(n) == 1:
                    # 只有项
                    if 'k' not in n[0]:
                        infor[m] = [infor[m - 1][0], infor[m - 1][1], infor[m][0]]
                    # 只有款
                    else:
                        infor[m] = [infor[m - 1][0], infor[m][0]]
                else:
                    infor[m] = [infor[m - 1][0], infor[m][0], infor[m][1]]
        if len(infor[0]) > 0:
            all_infor.append((fa, infor))
    return all_infor


def law_respilt(law_list):
    new_list = []
    for laws in law_list:
        if re.search('\|', laws) == None:
            new_list.append(laws)
            continue
        temp = re.split('[ |]', laws)
        temp = [x for x in temp if x != '']
        i = 0
        while i < len(temp):
            a = ''
            if i < len(temp) and re.search('t', temp[i]) != None:
                a += temp[i] + ' '
                i += 1
                while i < len(temp) and re.search('t', temp[i]) == None:
                    a += temp[i] + ' '
                    i += 1
            new_list.append(a)
    return new_list


def detect(law_list):
    for laws in law_list:
        # if re.search('\|',law)!=None:
        #     return  True
        if len(laws.split(" ")[:-1]) > 3:
            return True
    return False


def word_to_number(char):
    return {
        '一': 1,
        '二': 2,
        '三': 3,
        '四': 4,
        '五': 5,
        '六': 6,
        '七': 7,
        '八': 8,
        '九': 9,
        '十': 0,
        '、': '|',
    }.get(char, '')


def match(candidate, fa):  # 将法条和抽取结果作匹配
    flag=True
    for i in candidate:
        if re.search(i,fa) ==None:
            flag=False
            break
    return flag


def law_vector(law_list):
    key = law.keys()
    vec = list(np.zeros(len(LAW),dtype=int))
    for laws in law_list:
        if len(laws)==0:
            continue
        temp = laws.split(" ")
        temp = [x for x in temp if x != '']
        if temp[0] not in key:  # 不是属于贪污犯罪的法条
            continue
        candidate = law.get(temp[0])
        for tail in candidate:
            if match(tail,laws): #法条匹配
                vec[LAW.index(tail)]=1
    return vec


if __name__ == "__main__":
    wb = lw('纯贪污罪_一审_清洗.xlsx')
    ws = wb.worksheets[0]
    row = ws.max_row
    col = ws.max_column
    count = 0  # 计算数据为空的判决书个数
    count2 = 0  # 计算未能抽取法条的判决书个数
    # empty_result = []
    fire_write = openpyxl.Workbook()
    write = fire_write.create_sheet(index=0)
    o = 1
    # for i in range(1, col + 1):
    #     cell_value = ws.cell(row=1, column=i).value
    #     write.cell(row=1, column=i).value = cell_value
    num = 0
    for i in range(2, row + 1):
        # print(i)
        text = ws.cell(row=i, column=30).value
        if text == None:
            # empty_result.append(i)
            continue
        law_list = divide(catch(text))
        if len(law_list) == 0:
            # empty_result.append(i)
            continue
        law_com = []
        for j in law_list[0][1]:
            temp = ''
            for k in j:
                temp += k + ' '
            law_com.append(temp)
        law_com = law_respilt(law_com)

        vec = law_vector(law_com)
        if sum(vec) == 0:
            continue
        vec_str = ''.join(str(x) for x in vec)
        # if detect(law_com):
        #     print(text)
        #     print(law_com)
        #     print(vec_str)
        write.cell(row=o, column=1).value = i
        write.cell(row=o, column=2).value = ';'.join(law_com)
        write.cell(row=o, column=3).value = vec_str
        o += 1
    fire_write.save(filename="贪污受贿罪_一审_提取.xlsx")
