from sklearn.cluster import KMeans
from sklearn.cluster import *
from sklearn import metrics
import numpy as np
import openpyxl
import matplotlib.pylab as plt
import pandas as pd
from sklearn.cluster import Birch
from sklearn.metrics import calinski_harabaz_score

FILE_NAME = "E:\data_analysis\Judgement_information_extraction\database_xml\贪污受贿罪_一审_提取.xlsx"

file = FILE_NAME

def K_Means():
    df = pd.read_excel(file)

    # 提取判决法条
    # x = df.iloc[4]
    # print(x)
    # y = df.iloc[:,-1]
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    col = ws.max_column
    row = ws.max_row
    # row = 100
    data = []
    data_index = []
    for i in range(1, row + 1):
        data_index.append(int(ws.cell(row=i, column=1).value))
        data.append(list(str(ws.cell(row=i, column=3).value)))
    X = np.array(data)

    # ==============Ziyuan method====================
    # find out the suitable number of cluster
    # d = []
    # for i in range(1, 11):  # k取值1~11，做kmeans聚类，看不同k值对应的簇内误差平方和
    #     km = KMeans(n_clusters=i, init='k-means++', n_init=10, max_iter=300, random_state=0)
    #     km.fit(X)
    #     d.append(km.inertia_)
    # plt.axvline(x=4, linewidth=1.5, color='orangered', linestyle="--", label=':"knee" point')
    # plt.plot(range(1,11), d, marker='o')
    # plt.xlabel('number of clusters')
    # plt.ylabel('distortions')
    # plt.legend(loc='upper right', prop={'size': 8})
    # plt.show()
    #
    #
    # three_cluster_model = KMeans(n_clusters=4,  random_state=0)
    # three_cluster_model.fit(X)
    # label = list(three_cluster_model.labels_)
    # predict_model = three_cluster_model.predict(X)
    # # print(predict_model[4000:4200])
    # centres = three_cluster_model.cluster_centers_
    # # print(centres)
    # colors = ['r','c','b','y']
    # plt.figure
    # # print(X)
    # for j in range(4):
    #     index_set = np.where(predict_model == j)
    #     cluster = X[index_set]
    #     print(cluster)
    #     plt.scatter(cluster[:, 1], cluster[:, 2], c=colors[j], marker='.')
    #     plt.plot(centres[j][0], centres[j][1], 'o', markerfacecolor=colors[j], markeredgecolor='k',
    #              markersize=8)  # 画类别中心
    # plt.show()
    #=====Evaluation index=====


    # inertias = three_cluster_model.inertia_
    # adjust_rand_index = metrics.adjusted_rand_score((y_))

    # ===============Peijun method===================
    kmeans = KMeans(n_clusters=4, init='k-means++', n_init=10, max_iter=300, random_state=0).fit(X)
    label = list(kmeans.labels_)
    num = np.zeros(4, dtype=int)
    for i in label:
        num[i] += 1
    print(num)
    for i in range(1,row+1):
        ws.cell(row=i,column=4).value=label[i-1]
    wb.save(filename='贪污受贿罪_一审_提取.xlsx')


def spectral():
    # 提取判决法条
    # x = df.iloc[4]
    # print(x)
    # y = df.iloc[:,-1]
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    col = ws.max_column
    row = ws.max_row
    # row = 100
    data = []
    data_index = []
    for i in range(1, row + 1):
        data_index.append(int(ws.cell(row=i, column=1).value))
        temp = list(str(ws.cell(row=i, column=3).value))
        temp = [int(x) for x in temp]
        data.append(temp)
    X = np.array(data)
    # y_pre = Birch(n_clusters=None).fit_predict(X)
    label = Birch(n_clusters=4).fit_predict(X)
    num = np.zeros(4, dtype=int)
    for i in label:
        num[i] += 1
    print(num)
    for i in range(1, row + 1):
        ws.cell(row=i, column=4).value = label[i - 1]
    wb.save(filename='贪污受贿罪_一审_提取.xlsx')
    print("finish")

if __name__=="__main__":
    spectral()

